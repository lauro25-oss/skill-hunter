import uuid
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Request, status
from fastapi.responses import StreamingResponse, RedirectResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from slowapi import Limiter
from slowapi.util import get_remote_address
from app.database import get_db
from app.models.candidate import Candidate, CandidateStatus
from app.models.comment import CandidateComment
from app.schemas.candidate import CandidateOut, CandidateUpdate, CommentOut, CommentCreate
from app.services import parser, storage
from app.services.score import calcular_score
from app.auth import get_current_user

router = APIRouter(
    prefix="/candidates",
    tags=["Candidatos"],
    dependencies=[Depends(get_current_user)],
)
limiter = Limiter(key_func=get_remote_address)

ALLOWED_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}
MAX_SIZE_MB = 10


# ── Upload em lote ───────────────────────────────────────────

@router.post("/upload", status_code=status.HTTP_201_CREATED)
@limiter.limit("30/minute")
async def upload_curriculos(
    request: Request,
    files: list[UploadFile] = File(..., description="PDFs ou DOCX dos candidatos"),
    vaga_origem: str | None = Form(None, description="Nome da vaga para rastreamento"),
    db: AsyncSession = Depends(get_db),
):
    criados = []
    erros   = []

    for file in files:
        try:
            if file.content_type not in ALLOWED_TYPES:
                erros.append({"arquivo": file.filename, "erro": "Formato inválido. Use PDF ou DOCX."})
                continue

            file_bytes = await file.read()
            if len(file_bytes) > MAX_SIZE_MB * 1024 * 1024:
                erros.append({"arquivo": file.filename, "erro": f"Arquivo excede {MAX_SIZE_MB} MB."})
                continue

            data = await parser.parse_curriculum(file_bytes, file.filename)

            # Verificar duplicata: mesmo nome + (mesmo contato OU mesma experiência)
            _nome  = (data.get("nome") or "").strip()
            _email = (data.get("email") or "").strip().lower()
            _fone  = (data.get("telefone") or "").strip()
            _anos  = data.get("anos_experiencia")

            _dupes = (await db.execute(
                select(Candidate).where(
                    func.lower(func.trim(Candidate.nome)) == func.lower(_nome)
                )
            )).scalars().all()

            def _dup_match(d: Candidate) -> bool:
                email_igual = _email and d.email and _email == d.email.lower()
                fone_igual  = _fone  and d.telefone and _fone == d.telefone
                anos_igual  = (
                    _anos is not None
                    and d.anos_experiencia is not None
                    and abs(_anos - d.anos_experiencia) < 1
                )
                return bool(email_igual or fone_igual or anos_igual)

            if _dupes and any(_dup_match(d) for d in _dupes):
                erros.append({
                    "arquivo": file.filename,
                    "erro": (
                        f"Candidato '{_nome}' já existe no sistema "
                        f"(mesmo nome e dados coincidentes)."
                    ),
                })
                continue

            score = calcular_score(data)

            s3_key, b64_data = await storage.upload_file(
                file_bytes, file.filename, file.content_type or "application/pdf"
            )

            candidate = Candidate(
                id=uuid.uuid4(),
                nome=data["nome"],
                email=data.get("email"),
                telefone=data.get("telefone"),
                localizacao=data.get("localizacao"),
                cargo_atual=data.get("cargo_atual"),
                anos_experiencia=data.get("anos_experiencia"),
                experiencias=data.get("experiencias", []),
                hard_skills=data.get("hard_skills", []),
                resumo_ia=data.get("resumo_ia"),
                texto_bruto=data.get("texto_bruto"),
                vaga_origem=vaga_origem,
                gcs_url=s3_key or str(uuid.uuid4()),
                arquivo_s3_key=s3_key,
                arquivo_base64=b64_data,
                nome_arquivo=file.filename,
                score_aderencia=score,
            )
            db.add(candidate)
            await db.commit()
            await db.refresh(candidate)
            criados.append(CandidateOut.model_validate(candidate))

        except Exception as exc:
            await db.rollback()
            erros.append({"arquivo": file.filename, "erro": str(exc)})

    return {"criados": criados, "erros": erros}


# ── Stats ────────────────────────────────────────────────────

@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_db)):
    total_result = await db.execute(select(func.count(Candidate.id)))
    total = total_result.scalar() or 0

    rows = await db.execute(
        select(Candidate.status, func.count(Candidate.id)).group_by(Candidate.status)
    )
    counts = {row[0].value: row[1] for row in rows}

    score_result = await db.execute(
        select(func.avg(Candidate.score_aderencia)).where(Candidate.score_aderencia.isnot(None))
    )
    score_avg = round(float(score_result.scalar() or 0), 1)

    vagas_rows = await db.execute(
        select(Candidate.vaga_origem, func.count(Candidate.id))
        .where(Candidate.vaga_origem.isnot(None))
        .group_by(Candidate.vaga_origem)
        .order_by(func.count(Candidate.id).desc())
        .limit(6)
    )
    top_vagas = [{"vaga": r[0], "total": r[1]} for r in vagas_rows]

    approval_rows = await db.execute(
        select(Candidate.aprovado_cliente, func.count(Candidate.id))
        .where(Candidate.em_shortlist == True)
        .group_by(Candidate.aprovado_cliente)
    )
    approval = {"aprovado": 0, "reprovado": 0, "pendente": 0}
    for val, cnt in approval_rows:
        if val is True:
            approval["aprovado"] = cnt
        elif val is False:
            approval["reprovado"] = cnt
        else:
            approval["pendente"] = cnt

    timeline_raw = await db.execute(text("""
        SELECT
            TO_CHAR(DATE_TRUNC('month', criado_em), 'YYYY-MM') AS mes,
            COUNT(*) AS total
        FROM candidates
        WHERE criado_em >= NOW() - INTERVAL '6 months'
        GROUP BY DATE_TRUNC('month', criado_em)
        ORDER BY DATE_TRUNC('month', criado_em)
    """))
    timeline = [{"mes": row[0], "total": row[1]} for row in timeline_raw]

    return {
        "total":      total,
        "novo":       counts.get("novo", 0),
        "em_triagem": counts.get("em_triagem", 0),
        "shortlist":  counts.get("shortlist", 0),
        "aprovado":   counts.get("aprovado", 0),
        "rejeitado":  counts.get("rejeitado", 0),
        "contratado": counts.get("contratado", 0),
        "score_avg":  score_avg,
        "top_vagas":  top_vagas,
        "aprovacao":  approval,
        "timeline":   timeline,
    }


# ── GET único ────────────────────────────────────────────────

@router.get("/{candidate_id}", response_model=CandidateOut)
async def get_candidate(candidate_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    c = await _get_or_404(candidate_id, db)
    return CandidateOut.model_validate(c)


# ── Data URL / presigned URL do CV ───────────────────────────

@router.get("/{candidate_id}/cv-url")
async def get_cv_url(
    candidate_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    c = await _get_or_404(candidate_id, db)

    ext = (c.nome_arquivo or "").lower()
    content_type = (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        if ext.endswith(".docx") else "application/pdf"
    )
    filename = c.nome_arquivo or "curriculo.pdf"

    # R2/S3: backend baixa o arquivo e envia ao cliente (sem CORS)
    if c.arquivo_s3_key:
        try:
            file_bytes = await storage.get_file_bytes(c.arquivo_s3_key)
            return Response(
                content=file_bytes,
                media_type=content_type,
                headers={"Content-Disposition": f'inline; filename="{filename}"'},
            )
        except Exception:
            pass

    # Fallback: base64 no banco
    if c.arquivo_base64:
        file_bytes = storage.decode_file(c.arquivo_base64)
        return Response(
            content=file_bytes,
            media_type=content_type,
            headers={"Content-Disposition": f'inline; filename="{filename}"'},
        )

    raise HTTPException(404, "Este candidato não possui arquivo armazenado.")


# ── Re-parsear CV existente ──────────────────────────────────

@router.post("/{candidate_id}/reparse", response_model=CandidateOut)
async def reparse_candidate(candidate_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    c = await _get_or_404(candidate_id, db)

    if c.arquivo_s3_key:
        file_bytes = await storage.get_file_bytes(c.arquivo_s3_key)
    elif c.arquivo_base64:
        file_bytes = storage.decode_file(c.arquivo_base64)
    else:
        raise HTTPException(400, "Candidato sem arquivo armazenado para re-parsear.")

    data  = await parser.parse_curriculum(file_bytes, c.nome_arquivo or "curriculo.pdf")
    score = calcular_score(data)

    c.nome             = data["nome"]             or c.nome
    c.email            = data.get("email")        or c.email
    c.telefone         = data.get("telefone")     or c.telefone
    c.localizacao      = data.get("localizacao")  or c.localizacao
    c.cargo_atual      = data.get("cargo_atual")  or c.cargo_atual
    c.anos_experiencia = data.get("anos_experiencia") or c.anos_experiencia
    c.hard_skills      = data.get("hard_skills")  or c.hard_skills
    c.resumo_ia        = data.get("resumo_ia")    or c.resumo_ia
    c.texto_bruto      = data.get("texto_bruto")  or c.texto_bruto
    c.score_aderencia  = score

    await db.commit()
    await db.refresh(c)
    return CandidateOut.model_validate(c)


# ── PATCH ────────────────────────────────────────────────────

@router.patch("/{candidate_id}", response_model=CandidateOut)
async def update_candidate(
    candidate_id: uuid.UUID,
    payload: CandidateUpdate,
    db: AsyncSession = Depends(get_db),
):
    c = await _get_or_404(candidate_id, db)
    dados = payload.model_dump(exclude_unset=True)

    for field, value in dados.items():
        setattr(c, field, value)
    await db.commit()
    await db.refresh(c)

    return CandidateOut.model_validate(c)


# ── EXPORT EXCEL ─────────────────────────────────────────────

@router.get("/export/excel")
async def export_excel(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Candidate).order_by(Candidate.criado_em.desc())
    )
    candidates = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    header_fill  = PatternFill("solid", fgColor="1D4ED8")
    header_font  = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        bottom=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
    )
    alt_fill = PatternFill("solid", fgColor="EFF6FF")

    headers    = ["Nome", "Cargo Atual", "Status", "Localização", "Exp. (anos)", "Score",
                  "Hard Skills", "Shortlist", "Avaliação Cliente", "Vaga", "E-mail", "Telefone", "Adicionado"]
    col_widths = [28, 22, 14, 20, 12, 10, 40, 12, 18, 20, 28, 16, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    STATUS_LABELS = {
        "novo": "Novo", "em_triagem": "Em triagem", "shortlist": "Shortlist",
        "aprovado": "Aprovado", "rejeitado": "Rejeitado", "contratado": "Contratado",
    }

    for ri, c in enumerate(candidates, 2):
        aprovacao = ("Aprovado" if c.aprovado_cliente is True
                     else "Reprovado" if c.aprovado_cliente is False else "Pendente")
        row_data = [
            c.nome,
            c.cargo_atual or "",
            STATUS_LABELS.get(c.status.value if hasattr(c.status, "value") else str(c.status), str(c.status)),
            c.localizacao or "",
            c.anos_experiencia or "",
            round(c.score_aderencia) if c.score_aderencia is not None else "",
            "; ".join(c.hard_skills or []),
            "Sim" if c.em_shortlist else "Não",
            aprovacao,
            c.vaga_origem or "",
            c.email or "",
            c.telefone or "",
            c.criado_em.strftime("%d/%m/%Y") if c.criado_em else "",
        ]
        fill = alt_fill if ri % 2 == 0 else None
        for ci_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci_idx == 7))
            cell.border = thin_border
            if fill:
                cell.fill = fill
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"candidatos_skill_hunter_{__import__('datetime').date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── DELETE ───────────────────────────────────────────────────

@router.delete("/{candidate_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_candidate(candidate_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    c = await _get_or_404(candidate_id, db)
    if c.arquivo_s3_key:
        try:
            await storage.delete_file(c.arquivo_s3_key)
        except Exception:
            pass
    await db.delete(c)
    await db.commit()


# ── ANONIMIZAR (LGPD) ─────────────────────────────────────────

@router.post("/{candidate_id}/anonimizar", status_code=status.HTTP_200_OK)
async def anonimizar_candidato(candidate_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    c = await _get_or_404(candidate_id, db)
    if c.arquivo_s3_key:
        try:
            await storage.delete_file(c.arquivo_s3_key)
        except Exception:
            pass
    c.nome           = "[Dados removidos - LGPD]"
    c.email          = None
    c.telefone       = None
    c.localizacao    = None
    c.texto_bruto    = None
    c.resumo_ia      = None
    c.arquivo_base64 = None
    c.arquivo_s3_key = None
    c.nome_arquivo   = None
    c.gcs_url        = None
    c.notas          = None
    c.anonimizado    = True
    await db.commit()
    return {"ok": True, "mensagem": "Dados pessoais removidos conforme LGPD."}


# ── Comentários ──────────────────────────────────────────────

@router.get("/{candidate_id}/comments")
async def list_comments(candidate_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    await _get_or_404(candidate_id, db)
    result = await db.execute(
        select(CandidateComment)
        .where(CandidateComment.candidate_id == candidate_id)
        .order_by(CandidateComment.criado_em.asc())
    )
    return result.scalars().all()


@router.post("/{candidate_id}/comments", status_code=status.HTTP_201_CREATED)
async def add_comment(candidate_id: uuid.UUID, payload: CommentCreate, db: AsyncSession = Depends(get_db)):
    await _get_or_404(candidate_id, db)
    comment = CandidateComment(id=uuid.uuid4(), candidate_id=candidate_id, texto=payload.texto)
    db.add(comment)
    await db.commit()
    await db.refresh(comment)
    return CommentOut.model_validate(comment)


@router.delete("/{candidate_id}/comments/{comment_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_comment(candidate_id: uuid.UUID, comment_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(CandidateComment).where(
            CandidateComment.id == comment_id,
            CandidateComment.candidate_id == candidate_id,
        )
    )
    comment = result.scalar_one_or_none()
    if not comment:
        raise HTTPException(404, "Comentário não encontrado.")
    await db.delete(comment)
    await db.commit()


# ── Helper ───────────────────────────────────────────────────

async def _get_or_404(candidate_id: uuid.UUID, db: AsyncSession) -> Candidate:
    result = await db.execute(select(Candidate).where(Candidate.id == candidate_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Candidato não encontrado.")
    return c
